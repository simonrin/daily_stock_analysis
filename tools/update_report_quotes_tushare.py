#!/usr/bin/env python3
"""Update report market cells from Tushare.

The quote column intentionally omits trade date, turnover amount, and 60-day
price ranges. It shows latest close, monthly pct change, and PEG when available.
"""

from __future__ import annotations

import argparse
import html
import os
import re
from datetime import date, timedelta
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


TOKEN_KEYS = (
    "ASTOCKANA_TUSHARE_TOKEN",
    "TUSHARE_TOKEN",
    "TUSHARE_PRO_TOKEN",
    "TS_TOKEN",
    "TUSHARE_API_KEY",
)


def load_env(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values
    for raw in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")
    return values


def find_token(env_path: Path) -> str | None:
    dotenv = load_env(env_path)
    for key in TOKEN_KEYS:
        value = os.environ.get(key) or dotenv.get(key)
        if value:
            return value
    return None


def normalize_ts_code(value: object) -> str:
    raw = str(value or "").strip().upper()
    if not raw:
        return ""
    match = re.search(r"(\d{6})\.(SH|SZ)", raw)
    if match:
        return f"{match.group(1)}.{match.group(2)}"
    digits_match = re.search(r"\d{6}", raw)
    if not digits_match:
        return ""
    digits = digits_match.group(0)
    if digits.startswith(("600", "601", "603", "605", "688")):
        return f"{digits}.SH"
    return f"{digits}.SZ"


def stock_code_digits(value: object) -> str:
    match = re.search(r"\b(\d{6})(?:\.(?:SH|SZ))?\b", str(value or "").upper())
    return match.group(1) if match else ""


def ths_stock_url(value: object) -> str:
    digits = stock_code_digits(value)
    return f"https://stockpage.10jqka.com.cn/{digits}/" if digits else ""


def trade_dates(days_back: int = 45) -> Iterable[str]:
    today = date.today()
    for offset in range(days_back + 1):
        yield (today - timedelta(days=offset)).strftime("%Y%m%d")


def format_pct(value: float | None) -> str:
    return "待复核" if value is None else f"{value:+.1f}%"


def fetch_market_valuations(ts_codes: list[str], token: str) -> dict[str, str]:
    import tushare as ts

    pro = ts.pro_api(token)
    wanted = list(dict.fromkeys(ts_codes))
    closes: dict[str, float] = {}
    month_changes: dict[str, float] = {}
    quote_dates: dict[str, str] = {}

    end = date.today()
    start = end - timedelta(days=120)
    for code in wanted:
        try:
            frame = pro.daily(ts_code=code, start_date=start.strftime("%Y%m%d"), end_date=end.strftime("%Y%m%d"))
        except Exception:
            continue
        if frame is None or frame.empty:
            continue
        try:
            row = frame.sort_values("trade_date", ascending=False).iloc[0]
            closes[code] = float(row.get("close"))
            quote_dates[code] = str(row.get("trade_date", ""))
            recent = frame.sort_values("trade_date", ascending=False).reset_index(drop=True)
            if len(recent) > 20:
                prev_close = float(recent.iloc[20].get("close"))
                if prev_close > 0:
                    month_changes[code] = (closes[code] / prev_close - 1) * 100
        except Exception:
            continue

    pe_ttm: dict[str, float] = {}
    growth: dict[str, float] = {}

    for code in wanted:
        try:
            frame = pro.daily_basic(ts_code=code, trade_date=quote_dates.get(code, ""), fields="ts_code,trade_date,pe_ttm")
        except Exception:
            continue
        if frame is None or frame.empty:
            continue
        try:
            value = float(frame.iloc[0].get("pe_ttm"))
        except Exception:
            value = 0.0
        if value > 0:
            pe_ttm[code] = value

    for code in wanted:
        try:
            frame = pro.fina_indicator(ts_code=code, fields="ts_code,end_date,netprofit_yoy")
        except Exception:
            continue
        if frame is None or frame.empty:
            continue
        frame = frame.sort_values("end_date", ascending=False)
        for _, row in frame.iterrows():
            try:
                value = float(row.get("netprofit_yoy"))
            except Exception:
                continue
            if value > 0:
                growth[code] = value
                break

    market: dict[str, str] = {}
    for code in wanted:
        close = closes.get(code)
        quote = f"收盘 {close:.2f}，月涨跌幅 {format_pct(month_changes.get(code))}" if close else "行情待复核"
        pe = pe_ttm.get(code)
        yoy = growth.get(code)
        peg = f"PEG {pe / yoy:.2f}" if pe and yoy else "PEG 待复核"
        market[code] = f"{quote}；{peg}"
    return market


def locate_columns(ws) -> tuple[int, int]:
    header_values = [str(ws.cell(row=2, column=col_idx).value or "") for col_idx in range(1, ws.max_column + 1)]
    code_col = next((idx + 1 for idx, value in enumerate(header_values) if "股票代码" in value or "公司/股票代码" in value), 2)
    quote_col = next((idx + 1 for idx, value in enumerate(header_values) if "行情" in value or "PEG" in value or "估值" in value), 5)
    return code_col, quote_col


def normalize_layer_cell(value: object) -> str:
    text = str(value or "")
    colon = chr(0xFF1A)
    semi = chr(0xFF1B)
    if semi in text:
        left, right = text.split(semi, 1)
        if colon in left:
            left = left.split(colon, 1)[1]
        if colon in right:
            right = right.split(colon, 1)[1]
        return f"{left}/{right}"
    if colon in text:
        return text.split(colon, 1)[1]
    return text


def worksheet_rows(ws) -> tuple[list[str], list[list[str]]]:
    headers = [str(ws.cell(row=2, column=col_idx).value or "") for col_idx in range(1, ws.max_column + 1)]
    rows: list[list[str]] = []
    for row_idx in range(3, ws.max_row + 1):
        row = [str(ws.cell(row=row_idx, column=col_idx).value or "") for col_idx in range(1, ws.max_column + 1)]
        if any(cell.strip() for cell in row):
            rows.append(row)
    return headers, rows


def render_html_table(headers: list[str], rows: list[list[str]]) -> str:
    thead = "".join(f"<th>{html.escape(value)}</th>" for value in headers)
    tbody = []
    for row in rows:
        cells = []
        for idx, value in enumerate(row):
            cell = html.escape(value)
            url = ths_stock_url(value) if idx == 1 else ""
            if url:
                cell = f"<a href=\"{html.escape(url)}\">{cell}</a>"
            if idx in (0, 1, 4):
                cell = f"<strong>{cell}</strong>"
            cells.append(f"<td>{cell}</td>")
        tbody.append("<tr>" + "".join(cells) + "</tr>")
    return f"<table><thead><tr>{thead}</tr></thead><tbody>{''.join(tbody)}</tbody></table>"


def render_txt_table(headers: list[str], rows: list[list[str]]) -> str:
    return "\n".join(["\t".join(headers), *("\t".join(row) for row in rows)])


def update_sidecar(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    if not path.exists():
        return
    text = path.read_text(encoding="utf-8")
    text = text.replace("行情区间", "行情/PEG")
    text = text.replace("行情/估值", "行情/PEG")
    text = text.replace("行情（Tushare）", "行情/PEG")
    text = text.replace("行情/价格区间", "行情/PEG")
    text = text.replace("东方财富/交易所行情：表格候选公司的最新价、涨跌幅、成交额和估值分位。", "Tushare/交易所行情：收盘价、月涨跌幅和 PEG。")
    text = text.replace("Tushare/交易所行情：表格候选公司的收盘价和 PEG 比率。", "Tushare/交易所行情：收盘价、月涨跌幅和 PEG。")
    text = text.replace("Tushare/交易所行情：收盘价和 PEG 比率；不展示日期、成交额或近 60 交易日区间。", "Tushare/交易所行情：收盘价、月涨跌幅和 PEG。")
    text = text.replace("Tushare/交易所行情：收盘价、成交额和近 60 交易日区间。", "Tushare/交易所行情：收盘价、月涨跌幅和 PEG。")
    if path.suffix.lower() in {".html", ".htm"}:
        table = render_html_table(headers, rows)
        text = re.sub(r"<table>.*?</table>", table, text, count=1, flags=re.S)
    else:
        table = render_txt_table(headers, rows)
        marker = "核心候选公司横向对比表"
        start = text.find(marker)
        if start >= 0:
            table_start = text.find("\n", start)
            next_markers = [
                position
                for marker_text in ("\n\n过去", "\n\n需要", "\n\n本文")
                for position in [text.find(marker_text, table_start)]
                if position >= 0
            ]
            next_section = min(next_markers) if next_markers else len(text)
            if table_start >= 0 and next_section >= 0:
                text = text[:table_start + 1] + table + text[next_section:]
    path.write_text(text, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Update workbook quote/PEG cells from Tushare.")
    parser.add_argument("--workbook", required=True, help="Path to xlsx workbook.")
    parser.add_argument("--html", default=None, help="Optional HTML file to sync.")
    parser.add_argument("--txt", default=None, help="Optional text file to sync.")
    parser.add_argument("--env", default=".env", help="Path to .env file.")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    wb = load_workbook(workbook_path)
    ws = wb["Candidate Table"]
    code_col, quote_col = locate_columns(ws)

    codes_by_row: dict[int, str] = {}
    for row_idx in range(3, ws.max_row + 1):
        code = normalize_ts_code(ws.cell(row=row_idx, column=code_col).value)
        if code:
            codes_by_row[row_idx] = code
            link_cell = ws.cell(row=row_idx, column=code_col)
            link_cell.hyperlink = ths_stock_url(code)
            link_cell.font = Font(color="0563C1", underline="single")
    ws.cell(row=2, column=quote_col).value = "行情/PEG"
    for row_idx in range(3, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        cell.value = normalize_layer_cell(cell.value)
    for col_idx in (3, 7):
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical=cell.alignment.vertical or "top",
                text_rotation=cell.alignment.text_rotation,
                wrap_text=False,
                shrink_to_fit=cell.alignment.shrink_to_fit,
                indent=cell.alignment.indent,
            )

    token = find_token(Path(args.env))
    if token:
        valuations = fetch_market_valuations(list(codes_by_row.values()), token)
        missing_text = "行情待复核"
        status = "TUSHARE_MARKET_VALUATION_UPDATED"
    else:
        valuations = {}
        missing_text = "行情待复核"
        status = "TUSHARE_TOKEN_MISSING"

    quote_by_code: dict[str, str] = {}
    for row_idx, code in codes_by_row.items():
        value = valuations.get(code, f"{missing_text}；PEG 待复核")
        ws.cell(row=row_idx, column=quote_col).value = value
        quote_by_code[code] = value

    wb.save(workbook_path)
    headers, rows = worksheet_rows(ws)

    for sidecar in (args.html, args.txt):
        if sidecar:
            update_sidecar(Path(sidecar), headers, rows)

    print(f"{status} rows={len(codes_by_row)} workbook={workbook_path}")


if __name__ == "__main__":
    main()
