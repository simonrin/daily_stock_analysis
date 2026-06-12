#!/usr/bin/env python3
"""Generate and email the A-share supply-chain bottleneck weekly report.

This script is designed for GitHub Actions. Secrets are read from environment
variables only; do not commit .env files or real tokens to the repository.
"""

from __future__ import annotations

import html
import json
import mimetypes
import os
import smtplib
import ssl
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from email.header import Header
from email.message import EmailMessage
from email.utils import formataddr
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REPORT_DIR = Path("reports")
HEADERS = [
    "浜т笟閾惧眰绾?鍗＄偣鐜妭",
    "鍏徃/鑲＄エ浠ｇ爜",
    "鎺掑簭鍘熷洜",
    "涓€鍛ㄦ柊澧炶瘉鎹?,
    "琛屾儏鍖洪棿",
    "涓昏椋庨櫓",
    "寰呴獙璇佷簨瀹?鐮旂┒浼樺厛绾?,
]


@dataclass(frozen=True)
class Candidate:
    layer: str
    bottleneck: str
    company: str
    ts_code: str
    reason: str
    weekly_evidence: str
    evidence_strength: str
    risk: str
    validation: str
    priority: str


CANDIDATES: list[Candidate] = [
    Candidate("AI 鏁版嵁涓績鐢靛姏/娑插喎", "娑插喎娓╂帶銆佹満鎴跨儹绠＄悊", "鑻辩淮鍏?, "002837.SZ", "AI 鏈烘煖鍔熺巼瀵嗗害鎻愬崌锛屽喎鍗翠粠鎴愭湰椤瑰彉鎴愭墿浜х害鏉?, "AI 鏁版嵁涓績鐢靛姏銆佸喎鍗村拰鏈烘煖鍔熺巼瀵嗗害缁х画鏄墿瀹瑰墠缃害鏉?, "寮?, "娑插喎浠锋牸绔炰簤銆丄I 鏀跺叆鍗犳瘮涓嶆竻", "AI 鏁版嵁涓績璁㈠崟銆佹恫鍐锋敹鍏ュ崰姣斻€佹瘺鍒╃巼", "楂?),
    Candidate("AI 鏁版嵁涓績鐢靛姏/娑插喎", "绮惧瘑绌鸿皟銆佹恫鍐蜂笌鏁版嵁涓績鐜鎺у埗", "鐢宠彵鐜", "301018.SZ", "寮规€ч珮锛岃嫢璁㈠崟鍏戠幇鏇存晱鎰?, "姘?鐢?鍐峰嵈绾︽潫缁х画寮哄寲锛岄渶璺熻釜椤圭洰鍒惰鍗曞厬鐜?, "涓己", "浼板€兼尝鍔ㄣ€佽鍗曠‘璁ゆ粸鍚?, "澶у鎴烽」鐩笌浜や粯鍛ㄦ湡", "涓珮"),
    Candidate("鍥戒骇 AI 鑺墖/HBM/灏佽", "鍥戒骇 AI 鍔犻€熷櫒", "瀵掓绾?, "688256.SH", "鍥戒骇鏇夸唬鏀跨瓥鑻ユ帹杩涳紝绠楀姏鑺墖鏄渶纭害鏉熶箣涓€", "鍥戒骇绠楀姏鍩虹璁炬柦鎶曡祫棰勬湡浠嶅洿缁曡姱鐗囥€佸皝瑁呭拰鐢熸€佸睍寮€", "涓己", "鎬ц兘鐢熸€併€佸鎴烽泦涓€佷及鍊煎帇鍔?, "璁㈠崟銆佹敹鍏ョ‘璁ゃ€佽蒋浠剁敓鎬?, "楂?),
    Candidate("鍥戒骇 AI 鑺墖/HBM/灏佽", "鍥戒骇 CPU/GPU/鍔犻€熻绠楀钩鍙?, "娴峰厜淇℃伅", "688041.SH", "鍥戒骇绠楀姏鍩虹璁炬柦鏍稿績鍊欓€?, "鍥戒骇鑺墖渚涚粰鍜屽厛杩涘埗绋嬬害鏉熶粛鏄摱棰?, "涓己", "鍏堣繘鍒剁▼鍜屼緵搴旈摼闄愬埗", "鏁版嵁涓績瀹㈡埛銆佸簱瀛樺拰姣涘埄鐜?, "楂?),
    Candidate("鍏堣繘灏佽/HBM", "楂樼灏佹祴銆丆hiplet/鍏堣繘灏佽", "闀跨數绉戞妧", "600584.SH", "绠楀姏鑺墖鎵╀骇涓嶅彧鍗℃櫠鍦嗭紝涔熷崱灏佽銆佹祴璇曞拰鑹巼", "鍥戒骇 AI 鑺墖渚涚粰绾︽潫鎺ㄩ珮鍏堣繘灏佽閲嶈鎬?, "涓己", "浼犵粺灏佹祴鎷栫疮銆佽祫鏈紑鏀帇鍔?, "鍏堣繘灏佽鏀跺叆鍗犳瘮鍜屽鎴?, "涓珮"),
    Candidate("楂橀€熶簰杩?, "楂橀€熷厜妯″潡", "涓檯鏃垱", "300308.SZ", "AI 闆嗙兢甯﹀绾︽潫寮猴紝瀹㈡埛璁よ瘉鍜岃壇鐜囨瀯鎴愬鍨?, "AI 鏁版嵁涓績鎶曡祫棰勬湡缁х画鍗囨俯锛岄珮閫熶簰杩炰粛鏄牳蹇冨崱鐐?, "涓己", "娴峰瀹㈡埛闆嗕腑銆佷及鍊兼嫢鎸?, "800G/1.6T 璁㈠崟鍜屾瘺鍒╃巼", "楂?),
    Candidate("楂橀€熶簰杩?, "楂橀€熷厜妯″潡", "鏂版槗鐩?, "300502.SZ", "涓庢捣澶?AI 瀹㈡埛閾炬潯鐩稿叧搴﹂珮", "楂橀€熷厜妯″潡浠ｉ檯鍗囩骇鍜屽ぇ瀹㈡埛鎸佺画鎬т粛闇€楠岃瘉", "涓己", "璁㈠崟娉㈠姩銆佹妧鏈矾绾垮垏鎹?, "澶у鎴锋寔缁€с€佷骇鍝佷唬闄?, "楂?),
    Candidate("楂橀€?PCB/CCL", "AI 鏈嶅姟鍣?PCB", "娌數鑲′唤", "002463.SZ", "楂橀€熶簰杩炲悜 PCB/CCL 鎵╂暎锛屼綆鎹熻€楁潗鏂欏拰鑹巼鏄叧閿?, "AI 缃戠粶甯﹀鍜岄泦缇よ妯℃墿寮狅紝鍚戜綆鎹熻€?PCB/CCL 鎵╂暎", "涓?, "浼板€笺€佸鎴烽獙璇佽妭濂?, "AI 鏈嶅姟鍣ㄦ澘鏀跺叆鍗犳瘮", "涓珮"),
    Candidate("鍗婂浣撹澶?, "鍒昏殌銆佽杽鑶溿€佹竻娲楃瓑骞冲彴鍨嬭澶?, "鍖楁柟鍗庡垱", "002371.SZ", "鍥戒骇鏅跺渾鎵╀骇涓婇檺鍙栧喅浜庡叧閿澶囬獙璇佸拰浜や粯", "鍥戒骇 AI 鑺墖鍜屽厛杩涘皝瑁呴渶姹傚嚫鏄捐澶囧浗浜ф浛浠ｅ繀瑕佹€?, "寮?, "鏅跺渾鍘?capex 娉㈠姩銆佹牳蹇冮浂閮ㄤ欢", "鏂扮璁㈠崟銆侀獙鏀躲€佹瘺鍒╃巼", "楂?),
    Candidate("鍗婂浣撹澶?, "鍒昏殌銆丮OCVD", "涓井鍏徃", "688012.SH", "鍒昏殌宸ヨ壓楠岃瘉鍛ㄦ湡闀匡紝鏇夸唬闅惧害楂?, "鍏抽敭璁惧楠岃瘉鍛ㄦ湡闀匡紝浠嶆槸鎵╀骇鑺傚鐨勭‖绾︽潫", "寮?, "瀹㈡埛楠屾敹鍛ㄦ湡銆佷环鏍肩珵浜?, "鍒昏殌璁㈠崟鍜屽厛杩涜妭鐐归獙璇?, "楂?),
    Candidate("鍗婂浣撹澶?, "CVD/ALD 钖勮啘璁惧", "鎷撹崋绉戞妧", "688072.SH", "鍏堣繘鍒剁▼鍜屽厛杩涘皝瑁呴兘闇€瑕佽杽鑶滄矇绉兘鍔?, "钖勮啘娌夌Н璁惧鍙楀厛杩涘埗绋嬪拰鍏堣繘灏佽鍙岄噸鐗靛紩", "涓己", "鍗曚竴璧涢亾娉㈠姩銆佽鍗曢泦涓?, "鏂颁骇鍝佸鍏ュ拰楠屾敹", "涓珮"),
    Candidate("浜哄舰鏈哄櫒浜烘墽琛屽櫒", "璋愭尝鍑忛€熷櫒", "缁跨殑璋愭尝", "688017.SH", "濡傛灉鏈哄櫒浜鸿繘鍏ヨ妯¤瘯浜э紝绮惧瘑浼犲姩鍏堜簬鏁存満鍗′綇", "鏈哄櫒浜轰富棰橀渶鍥炲埌瀹㈡埛瀹氱偣銆佽壇鐜囧拰閲忎骇鑺傚楠岃瘉", "涓?, "閲忎骇鏃剁偣銆佷及鍊笺€佺珵浜?, "瀹㈡埛瀹氱偣銆侀噺浜ц鍗?, "涓珮"),
    Candidate("浜哄舰鏈哄櫒浜烘墽琛屽櫒", "涓濇潬/绮惧瘑闆堕儴浠?, "璐濇柉鐗?, "300580.SZ", "涓濇潬瀵垮懡鍜屼竴鑷存€у喅瀹氶噺浜у彲闈犳€?, "鏍蜂欢鍒伴噺浜х殑鍙潬鎬ч獙璇佹槸鏍稿績", "涓?, "鏍蜂欢鍒伴噺浜ц惤宸?, "涓濇潬璁㈠崟鍜岃壇鐜?, "涓?),
    Candidate("浜哄舰鏈哄櫒浜烘墽琛屽櫒", "鐢垫満/鎵ц鍣?, "楦ｅ織鐢靛櫒", "603728.SH", "绌哄績鏉?鎺у埗鐢垫満鏄叧鑺傝繍鍔ㄥ熀纭€", "鎵ц鍣ㄣ€佺數鏈哄拰鎺у埗绯荤粺浠嶉渶纭瀹㈡埛鍜岄噺浜ф敹鍏?, "涓?, "娴峰璁㈠崟涓嶉€忔槑銆侀渶姹傛尝鍔?, "鏈哄櫒浜烘敹鍏ュ崰姣?, "涓?),
    Candidate("绋€鍦熸案纾?, "楂樻€ц兘閽曢搧纭肩鏉?, "閲戝姏姘哥", "300748.SZ", "鏈哄櫒浜恒€佺數鍔ㄨ溅銆侀鐢靛叡鍚岄渶姹傦紝纾佹潗璁よ瘉鍛ㄦ湡闀?, "楂樼纾佹潗璁よ瘉鍛ㄦ湡鍜岀█鍦熶緵搴旂害鏉熶粛闇€璺熻釜", "涓己", "绋€鍦熶环鏍兼尝鍔ㄣ€佸嚭鍙ｆ斂绛栧彉鍖?, "璁㈠崟銆佷环鏍笺€佹捣澶栧鎴?, "涓珮"),
    Candidate("绋€鍦熸案纾?, "閽曢搧纭肩鏉?, "涓涓夌幆", "000970.SZ", "鑰佺墝纾佹潗渚涘簲鍟嗭紝鍙楃泭浜庨珮绔鏉愰渶姹?, "楂樼浜у搧鍗犳瘮鍐冲畾鐩堝埄寮规€?, "涓?, "鐩堝埄寮规€т笉纭畾", "楂樼浜у搧鍗犳瘮", "涓?),
    Candidate("鐢电綉涓昏澶?, "鎹㈡祦闃€銆佺户淇濄€佺洿娴佽緭鐢?, "璁哥户鐢垫皵", "000400.SZ", "鏂拌兘婧愬閫佸拰 AI 璐熻嵎閮介渶瑕佺數缃戞墿瀹?, "AI 璐熻嵎鍜屾柊鑳芥簮澶栭€佸己鍖栫數缃戜富璁惧闇€姹?, "寮?, "鎷涙爣鑺傚銆佸洖娆惧懆鏈?, "鍥界綉/鍗楃綉涓爣鍖?, "楂?),
    Candidate("鐢电綉涓昏澶?, "GIS/寮€鍏宠澶?, "骞抽珮鐢垫皵", "600312.SH", "鐗归珮鍘嬪拰涓荤綉鎶曡祫鍙€氳繃鎷涙爣楠岃瘉", "涓荤綉鎶曡祫鍜岀壒楂樺帇鎷涙爣鏄‖楠岃瘉璺緞", "寮?, "鎷涙爣浠锋牸涓嬮檷", "GIS 涓爣浠介鍜屾瘺鍒╃巼", "涓珮"),
    Candidate("鐢电綉鑷姩鍖?, "缁т繚銆佽皟搴︺€佽嚜鍔ㄥ寲", "鍥界數鍗楃憺", "600406.SH", "纭畾鎬у己锛屽脊鎬т綆浜庡皬甯傚€艰澶囧晢", "鏂板瀷鐢靛姏绯荤粺寤鸿闇€瑕佽皟搴︺€佺户淇濆拰鑷姩鍖?, "寮?, "浼板€煎脊鎬ф湁闄?, "鏂板瀷鐢靛姏绯荤粺璁㈠崟", "涓珮"),
    Candidate("浣庣┖鍩虹璁炬柦", "绌虹绯荤粺銆佷綆绌虹洃绠″钩鍙?, "鑾辨柉淇℃伅", "688631.SH", "浣庣┖缁忔祹鐪熷疄鍗＄偣鍦ㄧ┖鍩熺鐞嗗拰鐩戠", "浣庣┖缁忔祹鐪熷疄绾︽潫鍦ㄧ洃绠″钩鍙般€侀€氫俊瀵艰埅鍜岀┖鍩熺鐞?, "涓己", "鍦版柟璐㈡斂銆佽鍗曟姭闇蹭笉閫忔槑", "鍦版柟浣庣┖骞冲彴涓爣", "涓珮"),
    Candidate("浣庣┖鍩虹璁炬柦", "绌虹/闆疯揪/鍐涘伐鐢靛瓙", "鍥涘窛涔濇床", "000801.SZ", "浣庣┖鐩戠鍜岄€氫俊鐩戣闇€瑕佺‖浠跺熀纭€璁炬柦", "浣庣┖鐩戠闇€瑕侀€氫俊銆佺洃瑙嗗拰瀵艰埅鍩虹璁炬柦", "涓?, "涓婚灞炴€у己銆佷笟鍔℃媶鍒嗕笉娓?, "浣庣┖椤圭洰鏀跺叆鍗犳瘮", "涓?),
    Candidate("浣庣┖鍩虹璁炬柦", "閫氫俊瀵艰埅", "娴锋牸閫氫俊", "002465.SZ", "浣庣┖鏅鸿仈缃戦渶瑕侀€氫俊瀵艰埅鑳藉姏", "閫氫俊瀵艰埅椤圭洰涓爣鏄富瑕侀獙璇佺偣", "涓?, "璁㈠崟纭鎱?, "閫氫俊瀵艰埅椤圭洰涓爣", "涓?),
    Candidate("鏂拌兘婧愰珮绔潗鏂?, "鐢佃В娑?鏂板瀷鐢佃В璐?, "澶╄祼鏉愭枡", "002709.SZ", "鍥烘€?楂樺帇浣撶郴鑻ユ帹杩涳紝鏉愭枡楠岃瘉鏄墠缃害鏉?, "鍥烘€佸拰楂樺帇浣撶郴浠嶉渶鍏徃绾у鎴烽獙璇?, "涓?, "浼犵粺鐢佃В娑插懆鏈熶笅琛?, "鍥烘€佺數瑙ｈ川瀹㈡埛楠岃瘉", "涓?),
    Candidate("鏂拌兘婧愰珮绔潗鏂?, "楂樼闅旇啘", "鎭╂嵎鑲′唤", "002812.SZ", "楂樺畨鍏ㄩ殧鑶滀粛鏈夋妧鏈鍨掞紝浣嗚涓氫緵闇€鎵垮帇", "楂樼鑶滄潗鏂欐湁鎶€鏈鍨掞紝浣嗚涓氫环鏍煎帇鍔涗粛澶?, "涓?, "浠锋牸鎴樸€佷骇鑳借繃鍓?, "楂樼闅旇啘姣涘埄鐜囧拰瀹㈡埛", "涓?),
]

PUBLIC_SIGNAL_QUERIES = [
    "China AI data center power cooling liquid cooling",
    "China semiconductor equipment advanced packaging HBM",
    "China humanoid robot actuator harmonic reducer",
    "China ultra high voltage grid equipment tender",
    "China low altitude economy air traffic management",
    "China rare earth magnet export robotics",
]


def getenv(name: str, default: str = "") -> str:
    return os.getenv(name, default).strip()


def today_cn() -> date:
    return datetime.now(timezone(timedelta(hours=8))).date()


def trade_dates(days_back: int = 45) -> Iterable[str]:
    current = today_cn()
    for offset in range(days_back + 1):
        yield (current - timedelta(days=offset)).strftime("%Y%m%d")


def fetch_quotes(candidates: list[Candidate]) -> dict[str, str]:
    token = getenv("ASTOCKANA_TUSHARE_TOKEN") or getenv("TUSHARE_TOKEN")
    if not token:
        return {c.ts_code: "琛屾儏寰呭鏍? for c in candidates}

    try:
        import tushare as ts
    except Exception:
        return {c.ts_code: "琛屾儏寰呭鏍? for c in candidates}

    ts.set_token(token)
    pro = ts.pro_api(token)
    wanted = {c.ts_code for c in candidates}
    quotes: dict[str, str] = {}

    for trade_date in trade_dates():
        try:
            frame = pro.daily(trade_date=trade_date)
        except Exception:
            continue
        if frame is None or frame.empty:
            continue
        for _, row in frame.iterrows():
            code = str(row.get("ts_code", ""))
            if code not in wanted or code in quotes:
                continue
            close = float(row.get("close"))
            amount = row.get("amount")
            amount_text = ""
            if amount is not None:
                amount_text = f"锛屾垚浜ら {float(amount) / 100000:.2f} 浜垮厓"
            quotes[code] = f"{trade_date[4:6]}-{trade_date[6:8]} 鏀剁洏 {close:.2f}{amount_text}"
        if wanted.issubset(quotes):
            break

    ranges: dict[str, str] = {}
    end = today_cn()
    start = end - timedelta(days=120)
    for code in wanted:
        try:
            frame = pro.daily(ts_code=code, start_date=start.strftime("%Y%m%d"), end_date=end.strftime("%Y%m%d"))
        except Exception:
            continue
        if frame is None or frame.empty:
            continue
        recent = frame.sort_values("trade_date", ascending=False).head(60)
        ranges[code] = f"杩?0浜ゆ槗鏃ュ尯闂?{float(recent['low'].min()):.2f}-{float(recent['high'].max()):.2f}"

    out = {}
    for c in candidates:
        q = quotes.get(c.ts_code, "琛屾儏寰呭鏍?)
        r = ranges.get(c.ts_code)
        out[c.ts_code] = f"{q}锛泏r}" if r else q
    return out


def layer_summary() -> list[str]:
    seen: list[str] = []
    for c in CANDIDATES:
        if c.layer not in seen:
            seen.append(c.layer)
    return seen


def rows_for_report(quotes: dict[str, str]) -> list[list[str]]:
    rows = []
    for c in CANDIDATES:
        rows.append([
            f"浜т笟閾惧眰绾э細{c.layer}锛涘崱鐐圭幆鑺傦細{c.bottleneck}",
            f"{c.company}锛坽c.ts_code}锛?,
            c.reason,
            f"{c.weekly_evidence}锛涜瘉鎹己搴︼細{c.evidence_strength}",
            quotes.get(c.ts_code, "琛屾儏寰呭鏍?),
            c.risk,
            f"寰呴獙璇侊細{c.validation}锛涗紭鍏堢骇锛歿c.priority}",
        ])
    return rows


def fetch_public_signals(max_per_query: int = 2) -> list[dict[str, str]]:
    """Fetch recent public-source signals without requiring a paid API.

    GDELT is used only as a public article index. The report treats these as
    leads to verify, not as final proof of company fundamentals.
    """
    signals: list[dict[str, str]] = []
    for query in PUBLIC_SIGNAL_QUERIES:
        params = urllib.parse.urlencode({
            "query": query,
            "mode": "artlist",
            "format": "json",
            "maxrecords": str(max_per_query),
            "timespan": "7d",
            "sort": "hybridrel",
        })
        url = f"https://api.gdeltproject.org/api/v2/doc/doc?{params}"
        try:
            request = urllib.request.Request(url, headers={"User-Agent": "AStockAna-BottleneckWeekly/1.0"})
            with urllib.request.urlopen(request, timeout=12) as response:
                data = json.loads(response.read().decode("utf-8", errors="replace"))
        except Exception as exc:
            signals.append({"theme": query, "title": f"鍏紑鏉ユ簮寰呭鏍革細{exc.__class__.__name__}", "url": "", "date": ""})
            continue
        for article in data.get("articles", [])[:max_per_query]:
            title = str(article.get("title") or "").strip()
            article_url = str(article.get("url") or "").strip()
            seen_date = str(article.get("seendate") or "").strip()
            if title:
                signals.append({"theme": query, "title": title, "url": article_url, "date": seen_date})
    return signals[:12]


def write_xlsx(path: Path, rows: list[list[str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidate Table"
    ws.append(["A-Share Supply Chain Bottleneck Weekly"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    border = Border(
        left=Side(style="thin", color="9BB7D4"),
        right=Side(style="thin", color="9BB7D4"),
        top=Side(style="thin", color="9BB7D4"),
        bottom=Side(style="thin", color="9BB7D4"),
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{ws.max_row}"
    widths = [30, 20, 30, 34, 28, 24, 32]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws2 = wb.create_sheet("Layer Ranking")
    ws2.append(["鎺掑簭", "浜т笟閾惧崱鐐瑰眰绾?])
    for idx, layer in enumerate(layer_summary(), 1):
        ws2.append([idx, layer])
    for row in ws2.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 60
    wb.save(path)


def html_table(rows: list[list[str]]) -> str:
    header = "".join(f"<th>{html.escape(h)}</th>" for h in HEADERS)
    body = []
    for row in rows:
        body.append("<tr>" + "".join(f"<td>{html.escape(str(cell))}</td>" for cell in row) + "</tr>")
    return f"<table><thead><tr>{header}</tr></thead><tbody>{''.join(body)}</tbody></table>"


def render_html(rows: list[list[str]], signals: list[dict[str, str]]) -> str:
    layers = "".join(f"<li>{html.escape(layer)}</li>" for layer in layer_summary())
    table = html_table(rows)
    signal_items = "".join(
        "<li>"
        + html.escape(f"{item.get('date', '')} {item.get('theme', '')}: {item.get('title', '')}".strip())
        + (f" <a href=\"{html.escape(item['url'])}\">source</a>" if item.get("url") else "")
        + "</li>"
        for item in signals
    ) or "<li>鍏紑鏉ユ簮寰呭鏍搞€?/li>"
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <style>
    body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; color: #1f2937; line-height: 1.55; }}
    h2 {{ color: #174A7C; border-bottom: 1px solid #B8D0EA; padding-bottom: 8px; }}
    table {{ border-collapse: collapse; width: 100%; table-layout: fixed; }}
    th, td {{ border: 1px solid #9BB7D4; padding: 10px; vertical-align: top; word-break: break-word; }}
    th {{ background: #D9EAF7; font-weight: 700; }}
    tr:nth-child(even) td {{ background: #F7FBFF; }}
  </style>
</head>
<body>
  <h2>鏈懆鏈€鍊煎緱鐮旂┒鐨?{len(layer_summary())} 涓崱鐐瑰眰绾?/h2>
  <ol>{layers}</ol>
  <h2>鏍稿績鍊欓€夊叕鍙告í鍚戝姣旇〃</h2>
  {table}
  <h2>杩囧幓涓€鍛ㄥ叕寮€鏉ユ簮绾跨储</h2>
  <ol>{signal_items}</ol>
  <h2>闇€瑕佺户缁獙璇佺殑浜嬪疄娓呭崟</h2>
  <ol>
    <li>宸ㄦ疆璧勮鍜屼氦鏄撴墍鍏憡锛氶噸澶у悎鍚屻€佷腑鏍囥€佷骇鑳芥墿寤恒€佽皟鐮旂邯瑕佸拰椋庨櫓鎻愮ず銆?/li>
    <li>鍥界綉/鍗楃綉銆佸湴鏂逛綆绌虹粡娴庡钩鍙般€佹暟鎹腑蹇冮」鐩嫑鎶曟爣銆?/li>
    <li>Tushare/浜ゆ槗鎵€琛屾儏锛氭敹鐩樹环銆佹垚浜ら鍜岃繎 60 浜ゆ槗鏃ュ尯闂淬€?/li>
    <li>鍏徃绾ф敹鍏ョ粨鏋勶細AI銆佹満鍣ㄤ汉銆佷綆绌恒€佸厛杩涘皝瑁呯瓑涓氬姟鐪熷疄鍗犳瘮銆?/li>
  </ol>
  <h2>绠€鏄庣粨璁?/h2>
  <p>鏈懆鐮旂┒浼樺厛绾т粛搴斿厛鐪嬬湡瀹炰緵缁欑害鏉熷眰绾э紝鍐嶇湅涓偂寮规€с€侫I 鏁版嵁涓績鐢靛姏/娑插喎銆佸浗浜х畻鍔涜姱鐗囦笌鍏堣繘灏佽銆侀珮閫熶簰杩炪€佸崐瀵间綋璁惧銆佺數缃戜富璁惧鐨勮瘉鎹己搴﹁緝楂橈紱鏈哄櫒浜恒€佷綆绌虹粡娴庡拰鏂拌兘婧愰珮绔潗鏂欓渶瑕佹洿澶氳鍗曞拰鏀跺叆缁撴瀯楠岃瘉銆傛湰鏂囦粎涓虹爺绌朵紭鍏堢骇锛屼笉鏋勬垚涔板叆銆佸崠鍑烘垨鎸佹湁寤鸿銆?/p>
</body>
</html>
"""


def render_text(rows: list[list[str]], signals: list[dict[str, str]]) -> str:
    lines = [f"鏈懆鏈€鍊煎緱鐮旂┒鐨?{len(layer_summary())} 涓崱鐐瑰眰绾?]
    lines.extend(f"{idx}. {layer}" for idx, layer in enumerate(layer_summary(), 1))
    lines.append("")
    lines.append("鏍稿績鍊欓€夊叕鍙告í鍚戝姣旇〃")
    lines.append("\t".join(HEADERS))
    lines.extend("\t".join(row) for row in rows)
    lines.append("")
    lines.append("杩囧幓涓€鍛ㄥ叕寮€鏉ユ簮绾跨储")
    if signals:
        lines.extend(
            f"{idx}. {item.get('date', '')} {item.get('theme', '')}: {item.get('title', '')} {item.get('url', '')}".strip()
            for idx, item in enumerate(signals, 1)
        )
    else:
        lines.append("鍏紑鏉ユ簮寰呭鏍搞€?)
    lines.append("")
    lines.append("鏈枃浠呬负鐮旂┒浼樺厛绾э紝涓嶆瀯鎴愪拱鍏ャ€佸崠鍑烘垨鎸佹湁寤鸿銆?)
    return "\n".join(lines) + "\n"


def send_email(subject: str, text_body: str, html_body: str, attachments: list[Path]) -> None:
    if getenv("ASTOCKANA_DISABLE_EMAIL", "false").lower() == "true":
        print("EMAIL_DISABLED_BY_ASTOCKANA_DISABLE_EMAIL=true")
        return

    host = getenv("ASTOCKANA_SMTP_HOST")
    port = int(getenv("ASTOCKANA_SMTP_PORT", "465") or "465")
    user = getenv("ASTOCKANA_SMTP_USER")
    password = getenv("ASTOCKANA_SMTP_AUTH_CODE")
    recipient = getenv("ASTOCKANA_REPORT_RECIPIENT")
    from_name = getenv("ASTOCKANA_MAIL_FROM_NAME", "A-Share Supply Chain Weekly")
    missing = [name for name, value in {
        "ASTOCKANA_SMTP_HOST": host,
        "ASTOCKANA_SMTP_USER": user,
        "ASTOCKANA_SMTP_AUTH_CODE": password,
        "ASTOCKANA_REPORT_RECIPIENT": recipient,
    }.items() if not value]
    if missing:
        raise RuntimeError(f"Missing email secrets: {', '.join(missing)}")

    msg = EmailMessage()
    msg["From"] = formataddr((str(Header(from_name, "utf-8")), user))
    msg["To"] = recipient
    msg["Subject"] = subject
    msg.set_content(text_body, subtype="plain", charset="utf-8")
    msg.add_alternative(html_body, subtype="html", charset="utf-8")

    for path in attachments:
        ctype, _ = mimetypes.guess_type(path.name)
        if ctype is None:
            ctype = "application/octet-stream"
        maintype, subtype = ctype.split("/", 1)
        msg.add_attachment(path.read_bytes(), maintype=maintype, subtype=subtype, filename=path.name)

    with smtplib.SMTP_SSL(host, port, context=ssl.create_default_context(), timeout=30) as smtp:
        smtp.login(user, password)
        smtp.send_message(msg)
    print(f"EMAIL_SENT_TO={recipient}")


def main() -> None:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    report_date = today_cn().isoformat()
    prefix = REPORT_DIR / f"a-share-supply-chain-bottleneck-{report_date}"
    quotes = fetch_quotes(CANDIDATES)
    rows = rows_for_report(quotes)
    signals = fetch_public_signals()

    html_body = render_html(rows, signals)
    text_body = render_text(rows, signals)
    html_path = prefix.with_suffix(".html")
    text_path = prefix.with_suffix(".txt")
    doc_path = Path(str(prefix) + ".doc")
    xlsx_path = prefix.with_suffix(".xlsx")
    html_path.write_text(html_body, encoding="utf-8")
    text_path.write_text(text_body, encoding="utf-8")
    doc_path.write_text(html_body, encoding="utf-8")
    write_xlsx(xlsx_path, rows)

    subject_prefix = getenv("ASTOCKANA_MAIL_SUBJECT_PREFIX", "A-Share Supply Chain Bottleneck Weekly")
    subject = f"{subject_prefix} - {report_date}"
    send_email(subject, text_body, html_body, [xlsx_path, doc_path])


if __name__ == "__main__":
    main()

